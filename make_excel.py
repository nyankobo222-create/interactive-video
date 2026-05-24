from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# 共通スタイル
# ============================================================
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols_data, bg="2a9824", fg="FFFFFF"):
    for col, (val, width) in enumerate(cols_data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=11)
        c.alignment = align("center")
        c.border = border_all()
        ws.column_dimensions[get_column_letter(col)].width = width

def data_row(ws, row, values, bg="FFFFFF", bold=False, color="333333"):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=color)
        c.alignment = align()
        c.border = border_all()

def title_cell(ws, row, col, val, bg="E8F5E9", size=13):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg)
    c.font = font(bold=True, size=size, color="1B5E20")
    c.alignment = align("left")
    ws.row_dimensions[row].height = 26

def section(ws, row, val):
    c = ws.cell(row=row, column=1, value=val)
    c.fill = fill("F1F8E9")
    c.font = font(bold=True, size=11, color="2a9824")
    c.alignment = align("left")
    ws.row_dimensions[row].height = 22

ALT = "F9FBF9"

# ============================================================
# Sheet 1: システム全体構成
# ============================================================
ws1 = wb.active
ws1.title = "①システム全体構成"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 32

title_cell(ws1, 1, 1, "インタラクティブ動画システム　全体構成図", size=14)
ws1.merge_cells("A1:E1")

# ブロック図（テキストで）
rows_diagram = [
    (3,  "【ユーザー（スマホ/PC）】"),
    (4,  "ブラウザで URL を開く"),
    (5,  "↓"),
    (6,  "【Railway（Webサーバー）】"),
    (7,  "HTML / CSS / JavaScript を配信"),
    (8,  "（プレーヤーのコード一式）"),
    (9,  "↓ 動画ファイルのURLを参照"),
    (10, "【Cloudflare R2（動画ストレージ）】"),
    (11, "C01.mp4, C02.mp4 ... C14.mp4 を配信"),
]
for r, val in rows_diagram:
    c = ws1.cell(row=r, column=1, value=val)
    if val.startswith("【"):
        c.fill = fill("2a9824")
        c.font = font(bold=True, color="FFFFFF", size=11)
    elif val == "↓" or val == "↓ 動画ファイルのURLを参照":
        c.font = font(size=14, color="2a9824")
        c.alignment = align("center")
    else:
        c.fill = fill("F1F8E9")
        c.font = font(size=10, color="333333")
    c.alignment = align("left")
    ws1.row_dimensions[r].height = 22
ws1.column_dimensions["A"].width = 45

# 右側に役割説明
header_row(ws1, 3, [
    ("場所", 18), ("役割", 30), ("具体的な中身", 40), ("費用目安", 16), ("備考", 24)
], bg="37474F")

infra = [
    ("ユーザー端末",   "動画を視聴・操作する",         "スマホ / PC のブラウザ",            "無料", ""),
    ("Railway",       "Webサーバー（コード配信）",     "HTML, CSS, JS（プレーヤー本体）",    "無料〜$5/月", "GitHub と連携して自動デプロイ"),
    ("Cloudflare R2", "動画ファイルの保管・配信",      "MP4ファイル（C01〜C14）",            "無料（10GB/月まで）", "CDN配信で世界中から高速アクセス"),
    ("GitHub",        "コードのバージョン管理",        "ソースコード一式",                   "無料", "Railwayと連携"),
]
for i, row in enumerate(infra):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    data_row(ws1, 4 + i, row, bg=bg)

ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 40
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 24  # unused but set

# ============================================================
# Sheet 2: ファイル構成と役割
# ============================================================
ws2 = wb.create_sheet("②ファイル構成と役割")
ws2.sheet_view.showGridLines = False
ws2.row_dimensions[1].height = 32
title_cell(ws2, 1, 1, "ファイル構成と各ファイルの役割", size=14)
ws2.merge_cells("A1:F1")

header_row(ws2, 2, [
    ("ファイルパス", 42),
    ("種類", 14),
    ("役割", 36),
    ("触る頻度", 14),
    ("誰が触るか", 16),
    ("補足", 36),
], bg="2a9824")

files = [
    ("src/configs/fken.js",               "設定",   "動画URLとボタンラベルを定義。クライアントごとに1ファイル作る",        "★★★よく触る", "制作者",       "新クライアントが来たらこれをコピーして編集するだけ"),
    ("src/components/InteractivePlayer.jsx","ロジック","プレーヤーの状態管理（再生・停止・チャプター切替）",              "ほぼ触らない", "エンジニア",   "どのクライアントでも共通で使う「エンジン」"),
    ("src/components/BranchMenu.jsx",      "UI",     "ブランチ選択オーバーレイ（大きなボタンが3つ並ぶ画面）",             "たまに触る",   "エンジニア",   "デザイン変更時に編集"),
    ("src/components/BottomBar.jsx",       "UI",     "動画再生中に下部に表示される小ボタン群",                           "たまに触る",   "エンジニア",   "デザイン変更時に編集"),
    ("src/components/DemoChapter.jsx",     "UI",     "動画URLがない場合に表示するデモ用プレースホルダー",                 "ほぼ触らない", "エンジニア",   "開発・提案時のデモ用"),
    ("src/components/InteractivePlayer.css","スタイル","全コンポーネントの見た目（色・サイズ・アニメーション）",           "たまに触る",   "エンジニア",   "配色やフォントを変えるとき"),
    ("server.js",                          "サーバー","Railway上でビルド済みファイルを配信するNode.jsサーバー",           "ほぼ触らない", "エンジニア",   ""),
    ("package.json",                       "設定",   "使用ライブラリの一覧（React, Vite, Express等）",                  "ほぼ触らない", "エンジニア",   ""),
    ("nixpacks.toml",                      "設定",   "Railwayへのデプロイ手順を定義",                                   "ほぼ触らない", "エンジニア",   ""),
    ("C01.mp4 〜 C14.mp4",                 "動画",   "各チャプターの動画ファイル（Cloudflare R2に保存）",                 "★★★よく触る", "制作者",       "編集済みMP4をR2にアップロードするだけ"),
]
for i, row in enumerate(files):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    bold = row[3] == "★★★よく触る"
    data_row(ws2, 3 + i, row, bg=bg, bold=bold)

for col, w in zip("ABCDEF", [42, 14, 36, 14, 16, 36]):
    ws2.column_dimensions[get_column_letter(ord(col)-64)].width = w

# ============================================================
# Sheet 3: 動作フロー
# ============================================================
ws3 = wb.create_sheet("③動作フロー")
ws3.sheet_view.showGridLines = False
ws3.row_dimensions[1].height = 32
title_cell(ws3, 1, 1, "プレーヤーの動作フロー（FKENの場合）", size=14)
ws3.merge_cells("A1:E1")

header_row(ws3, 2, [
    ("ステップ", 8),
    ("状態名", 20),
    ("画面の様子", 36),
    ("ユーザーの操作", 24),
    ("次の状態", 20),
], bg="2a9824")

flows = [
    ("1", "intro（イントロ）",
     "C01動画が自動再生。ナレーション+映像が流れる",
     "（自動）",
     "15秒経過→branch_select"),
    ("2", "branch_select（選択待ち）",
     "C01が停止し、3つのボタンが大きく表示される（緑色・パルスアニメ付き）\n電話・問い合わせリンクも表示",
     "ボタンをタップ（01/02/03）",
     "branch_playing"),
    ("3", "branch_playing（再生中）",
     "選択したブランチのチャプターが順番に自動再生\n（例：01を選んだ→C02→C03→C04→C05）\n画面下部にボトムバーが常時表示",
     "ボトムバーのボタンで別ブランチへ切替も可能\n電話・メールリンクをタップ",
     "全チャプター終了→end_menu"),
    ("4", "end_menu（エンド）",
     "C14が再生され終了後、3つのボタン＋「トップへ戻る」が表示\n他のブランチを見るよう促す",
     "別のブランチを選ぶ / トップへ戻る",
     "branch_playing / intro"),
]
for i, row in enumerate(flows):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    data_row(ws3, 3 + i, row, bg=bg)
    ws3.row_dimensions[3 + i].height = 60

for col, w in zip([1,2,3,4,5], [8, 20, 36, 24, 20]):
    ws3.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# Sheet 4: UIレイヤーの仕組み
# ============================================================
ws4 = wb.create_sheet("④UIレイヤーの仕組み")
ws4.sheet_view.showGridLines = False
ws4.row_dimensions[1].height = 32
title_cell(ws4, 1, 1, "UIレイヤー（動画の上にボタンを重ねる仕組み）", size=14)
ws4.merge_cells("A1:D1")

# 説明
for r, (val, bg_c, bold) in enumerate([
    ("【たとえ話】", "E8F5E9", True),
    ("スライドショーのスライドに「付箋」を貼るイメージ。", "FFFFFF", False),
    ("スライド（動画）はそのまま流れていて、付箋（ボタン）が上に浮いているだけ。", "FFFFFF", False),
    ("", "FFFFFF", False),
    ("【技術的な仕組み】CSS の position プロパティを使う", "E8F5E9", True),
    ("position: relative → 「この箱を基準にする」と宣言（プレーヤーの外枠）", "FFFFFF", False),
    ("position: absolute → 「基準の箱の上に自由に配置する」（ボタンのUI）", "FFFFFF", False),
    ("inset: 0           → 「上下左右すべて0＝全面を覆う」", "FFFFFF", False),
    ("", "FFFFFF", False),
    ("【重ね順のイメージ】", "E8F5E9", True),
], 1):
    c = ws4.cell(row=2 + r, column=1, value=val)
    c.fill = fill(bg_c)
    c.font = font(bold=bold, size=11)
    c.alignment = align("left", wrap=False)
    ws4.row_dimensions[2 + r].height = 20
ws4.column_dimensions["A"].width = 70

header_row(ws4, 13, [
    ("層（上が手前）", 16), ("HTML要素", 20), ("CSSキーワード", 24), ("役割", 30)
], bg="2a9824")

layers = [
    ("3層目（最前面）", "<div class='branch-menu'>",  "position: absolute / inset: 0 / z-index: 10",  "ブランチ選択UI（タップで反応）"),
    ("2層目",          "<div class='bottom-bar'>",    "position: absolute / bottom: 0",               "ボトムナビバー"),
    ("1層目（最背面）", "<video>",                    "width: 100% / height: 100%",                    "動画本体（ただ再生するだけ）"),
    ("外枠",           "<div class='player'>",        "position: relative（基準点）",                  "全レイヤーを包む箱"),
]
for i, row in enumerate(layers):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    data_row(ws4, 14 + i, row, bg=bg)

for col, w in zip([1,2,3,4], [16, 20, 24, 30]):
    ws4.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# Sheet 5: 設定ファイルの読み方
# ============================================================
ws5 = wb.create_sheet("⑤設定ファイルの読み方")
ws5.sheet_view.showGridLines = False
ws5.row_dimensions[1].height = 32
title_cell(ws5, 1, 1, "設定ファイル（fken.js）の読み方", size=14)
ws5.merge_cells("A1:D1")

header_row(ws5, 2, [
    ("設定項目", 22), ("書いてあること", 30), ("プレーヤーへの影響", 36), ("変更例", 28)
], bg="2a9824")

settings = [
    ("company.name",      "株式会社FKEN",                       "エンドメニューに表示される会社名",                   "→ 別会社名に変えるだけ"),
    ("company.phone",     "0833-57-4497",                       "電話ボタンのリンク先番号",                           "→ 別の番号に変えるだけ"),
    ("company.contactUrl","https://fken.net/#anker-contact",    "問い合わせボタンのリンク先URL",                      "→ 別のURLに変えるだけ"),
    ("theme.primary",     "#2a9824（緑色）",                    "全ボタンの色・プログレスバーの色",                    "→ #E53935 にすれば赤になる"),
    ("chapters.C01.url",  "null（デモ）またはMP4のURL",          "C01として再生する動画ファイルを指定",                 "→ R2のURLを入れると本番動画になる"),
    ("chapters.C01.demoDuration", "15（秒）",                  "デモモード時にC01が何秒で終わるか",                   "→ 実際の動画を入れたら不要"),
    ("flow.intro.pauseAt","15（秒）",                           "C01動画を何秒で停止してメニューを出すか",             "→ 動画の尺に合わせて変更"),
    ("flow.branches",     "3つのブランチ定義",                  "ボタンのラベルと再生するチャプターの順番を決める",     "→ ブランチを2つにも4つにもできる"),
    ("flow.endMenu",      "C14",                                "全ブランチ終了後に再生するチャプターID",               "→ エンド動画を変える場合に変更"),
]
for i, row in enumerate(settings):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    data_row(ws5, 3 + i, row, bg=bg)

for col, w in zip([1,2,3,4], [22, 30, 36, 28]):
    ws5.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# Sheet 6: 新クライアント対応手順
# ============================================================
ws6 = wb.create_sheet("⑥新クライアント対応手順")
ws6.sheet_view.showGridLines = False
ws6.row_dimensions[1].height = 32
title_cell(ws6, 1, 1, "新しいクライアントに納品するときの手順", size=14)
ws6.merge_cells("A1:E1")

header_row(ws6, 2, [
    ("手順", 6), ("作業内容", 28), ("誰がやる", 14), ("所要時間目安", 14), ("補足", 36)
], bg="2a9824")

steps = [
    ("1", "クライアントから動画素材を受け取る（撮影済みMP4）",                 "クライアント",  "－",        "C01〜C14 に対応するファイルが揃っていればOK"),
    ("2", "Cloudflare R2 に新しいフォルダを作成\n例：r2.dev/abc-company/",     "制作者",       "10分",      "R2のダッシュボードからフォルダを作るだけ"),
    ("3", "動画ファイルをR2にアップロード\n（C01.mp4, C02.mp4 … C14.mp4）",    "制作者",       "30分〜1時間","ファイルサイズによる。ドラッグ&ドロップでOK"),
    ("4", "fken.js をコピーして新しい設定ファイルを作成\n例：abc_company.js",   "制作者",       "15分",      "会社名・電話番号・URL・ボタンラベルを変える"),
    ("5", "各chapterのurlにR2のURLを貼り付ける",                               "制作者",       "10分",      "コピペ作業"),
    ("6", "App.jsxで読み込む設定ファイルを切り替える\n（またはURLパラメータで動的切替）","エンジニア","5分",   "将来的には ?client=fken のようなURL設計も可能"),
    ("7", "ローカルで動作確認（npm run dev）",                                  "エンジニア",   "15分",      "全ブランチを一通り操作して確認"),
    ("8", "GitHubにpush → Railwayが自動デプロイ",                              "エンジニア",   "5分",       "pushするだけで自動的に本番反映"),
    ("9", "クライアントにURLを共有して納品",                                    "制作者",       "5分",       "例：https://fken.yourapp.com"),
]
for i, row in enumerate(steps):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    data_row(ws6, 3 + i, row, bg=bg)
    ws6.row_dimensions[3 + i].height = 40

for col, w in zip([1,2,3,4,5], [6, 28, 14, 14, 36]):
    ws6.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# Sheet 7: コスト比較
# ============================================================
ws7 = wb.create_sheet("⑦コスト比較")
ws7.sheet_view.showGridLines = False
ws7.row_dimensions[1].height = 32
title_cell(ws7, 1, 1, "コスト比較（ミル vs 自作）", size=14)
ws7.merge_cells("A1:D1")

header_row(ws7, 2, [
    ("項目", 24), ("ミル（既存サービス）", 24), ("自作システム", 24), ("備考", 30)
], bg="2a9824")

costs = [
    ("初期費用",         "高（プラン料金）",          "開発工数のみ（今回は構築済み）", ""),
    ("月額（1クライアント）","数万円/月（目安）",       "ほぼ0円",                      "R2は10GB/月まで無料"),
    ("月額（5クライアント）","数万円×5",               "変わらずほぼ0円",               "追加しても費用はほぼ増えない"),
    ("動画ストレージ",    "プランに含む（上限あり）",   "Cloudflare R2（10GB無料）",     "超えても$0.015/GB"),
    ("カスタマイズ",      "制限あり（テンプレ範囲）",   "完全自由",                      "ボタン数・デザイン・分岐ロジック等"),
    ("独自ドメイン",      "オプション（有料の場合も）", "Railway で無料設定可能",         ""),
    ("データ所有",        "サービス側",                "自分（GitHubで管理）",           "サービス終了リスクなし"),
    ("技術的な手間",      "ほぼなし",                  "初期構築のみ（構築済み）",        "追加クライアントは設定変更のみ"),
]
for i, row in enumerate(costs):
    bg = "FFFFFF" if i % 2 == 0 else ALT
    bold = i == 1 or i == 2
    data_row(ws7, 3 + i, row, bg=bg, bold=bold)

for col, w in zip([1,2,3,4], [24, 24, 24, 30]):
    ws7.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# 保存
# ============================================================
output_path = r"D:\interactive-video\インタラクティブ動画システム_構成説明.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
